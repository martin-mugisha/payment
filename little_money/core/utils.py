from django.utils.translation import gettext as _
import io
import datetime
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib.units import inch
from reportlab.lib import colors
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO

def is_admin(user):
    return hasattr(user, 'role') and user.role == 'admin'

def is_staff(user):
    return hasattr(user, 'role') and user.role == 'staff'

def is_client(user):
    return hasattr(user, 'role') and user.role == 'client'

def format_phone_number(number: str) -> str:
    """
    Format the phone number to the specified MTN/Airtel format:
    075XXXXXXX, 070XXXXXXX, 074XXXXXXX, 076XXXXXXX, 078XXXXXXX, 077XXXXXXX, 079XXXXXXX
    Assumes input number is a string of digits possibly with country code.
    """
    if not number:
        return ''
    # Remove non-digit characters
    digits = ''.join(filter(str.isdigit, number))
    # Remove country code if present (e.g., 256)
    if digits.startswith('256'):
        digits = digits[3:]
    elif digits.startswith('0'):
        digits = digits[1:]
    # Now digits should be 9 digits
    if len(digits) == 9:
        # Format as 07Xxxxxxxx
        return '0' + digits
    else:
        # Return original if format unknown
        return number
def generate_receipt_pdf(transaction):
    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    # --- Header ---
    header_height = 70
    p.setFillColorRGB(0.09, 0.28, 0.47)  # Deep blue
    p.rect(0, height - header_height, width, header_height, fill=1, stroke=0)

    p.setFont("Helvetica-Bold", 18)
    p.setFillColor(colors.white)
    p.drawString(inch, height - header_height + 20, "ManguPay - Payment Receipt")

    # --- Receipt Box ---
    box_top = height - inch - 50
    box_bottom = inch + 100
    p.setFillColorRGB(0.97, 0.97, 0.97)  # Light grey background
    p.roundRect(inch, box_bottom, width - 2 * inch, box_top - box_bottom, 10, fill=1, stroke=0)

    p.setStrokeColor(colors.black)
    p.setLineWidth(0.5)
    p.roundRect(inch, box_bottom, width - 2 * inch, box_top - box_bottom, 10, fill=0, stroke=1)

    # --- Transaction Info ---
    p.setFont("Helvetica-Bold", 12)
    p.setFillColor(colors.black)
    p.drawString(inch + 20, box_top - 30, "Transaction Details")

    p.setFont("Helvetica", 10)
    line_gap = 18
    y = box_top - 60

    # Combine date and time if needed
    try:
        timestamp = datetime.datetime.combine(transaction.date, transaction.time)
    except Exception:
        timestamp = None

    fields = [
        ("Transaction ID", transaction.transaction_id or "N/A"),
        ("Date", transaction.date.strftime("%Y-%m-%d") if transaction.date else "N/A"),
        ("Time", transaction.time.strftime("%H:%M:%S") if transaction.time else "N/A"),
        ("Recipient", transaction.recipient or "N/A"),
        ("Phone", transaction.phone or "N/A"),
        ("Amount", f"UGX {transaction.amount:,.0f}"),
        ("Transaction Type", transaction.transaction_type or "N/A"),
        ("Channel", transaction.payment_method or "N/A"),
        ("Status", transaction.status or "N/A"),
        ("Description", transaction.description or "None"),
    ]

    for label, value in fields:
        p.drawString(inch + 40, y, f"{label}:")
        p.drawRightString(width - inch - 40, y, str(value))
        y -= line_gap

    # --- Signature Section ---
    signature_y = box_bottom - 30
    p.setFont("Helvetica", 10)
    p.drawString(inch, signature_y, "Signed by:")
    p.setDash(1, 2)
    p.line(inch + 60, signature_y, inch + 260, signature_y)
    p.setDash()

    # --- Stamp Box ---
    stamp_width = 100
    stamp_height = 50
    p.setStrokeColor(colors.gray)
    p.rect(width - inch - stamp_width, signature_y - stamp_height + 10, stamp_width, stamp_height, fill=0)
    p.setFont("Helvetica-Oblique", 8)
    p.drawCentredString(width - inch - stamp_width / 2, signature_y - stamp_height / 2 + 10, "Stamp here")

    # --- Footer ---
    footer_y = 0.4 * inch
    p.setFont("Helvetica-Oblique", 8)
    p.setFillColor(colors.gray)
    p.drawCentredString(width / 2, footer_y + 12, "Thank you for choosing ManguPay")
    p.drawCentredString(width / 2, footer_y, f"Generated on {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    p.showPage()
    p.save()

    buffer.seek(0)
    return buffer

def generate_statement_xlsx(transactions, lang='en'):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = _("Transactions")

    headers = [
        _('Transaction ID'), _('Date'), _('Time'), _('Recipient'), _('Phone Number'),
        _('Amount'), _('Transaction Type'), _('Status'), _('Channel'), _('Description')
    ]
    ws.append(headers)

    # Bold headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = Font(bold=True)

    total_amount = 0
    for txn in transactions:
        row = [
            txn.transaction_id,
            txn.date.strftime("%Y-%m-%d") if txn.date else '',
            txn.time.strftime("%H:%M:%S") if txn.time else '',
            txn.recipient or '',
            txn.phone or '',
            float(txn.amount),
            txn.transaction_type or '',
            txn.status or '',
            txn.payment_method or '',
            txn.description or '',
        ]
        ws.append(row)
        total_amount += float(txn.amount)

        # Conditional formatting based on status
        status = txn.status.lower() if txn.status else ''
        status_fill = None
        if "success" in status:
            status_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
        elif "fail" in status or "error" in status:
            status_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light red

        if status_fill:
            status_cell = ws.cell(row=ws.max_row, column=8)
            status_cell.fill = status_fill

    # Add totals row
    total_label_cell = ws.cell(row=ws.max_row + 2, column=5)
    total_label_cell.value = _("Total Amount")
    total_label_cell.font = Font(bold=True)

    total_value_cell = ws.cell(row=ws.max_row, column=6)
    total_value_cell.value = total_amount
    total_value_cell.font = Font(bold=True)

    # Auto column width
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(max_length + 2, 15)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output